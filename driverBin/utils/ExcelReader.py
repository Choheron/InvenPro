from tkinter import Tk
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.workbook.workbook import Workbook
from datetime import date, datetime
import os
import sys
import json
import tkinter as tk

def readCPUlist():
    CPUlistFilepath = (os.getcwd() + '/../../../MASLD COMPUTER LOGS.xlsx')
    cpuBook = load_workbook(CPUlistFilepath, data_only = True)
    ws = cpuBook.active

    outDict = {}
    cRow = 3 # Value based off of first occurance of data in MASTER COMPUTER EXCEL SHEET
    CPUCount = 0
    while ws[f'B{cRow}'].value:
        cpuDict = {}
        cpuDict['idnum'] = ws[f'B{cRow}'].value
        cpuDict['ipv4'] = ws[f'C{cRow}'].value
        cpuDict['location'] = ws[f'D{cRow}'].value
        cpuDict['userInit'] = ws[f'E{cRow}'].value
        cpuDict['user'] = ws[f'F{cRow}'].value
        cpuDict['build'] = ws[f'G{cRow}'].value
        cpuDict['type'] = ws[f'H{cRow}'].value
        cpuDict['cost'] = ws[f'I{cRow}'].value
        cpuDict['year'] = ws[f'J{cRow}'].value
        cpuDict['manufacturer'] = ws[f'K{cRow}'].value
        cpuDict['softwares'] = ws[f'L{cRow}'].value
        cpuDict['wifi'] = ws[f'M{cRow}'].value
        cpuDict['gpu'] = ws[f'N{cRow}'].value
        cpuDict['cpu'] = ws[f'O{cRow}'].value
        cpuDict['mobo'] = ws[f'P{cRow}'].value
        cpuDict['ram'] = ws[f'Q{cRow}'].value
        cpuDict['storage'] = ws[f'R{cRow}'].value
        cpuDict['os'] = ws[f'S{cRow}'].value
        cpuDict['lastRowUpdate'] = ws[f'T{cRow}'].value

        outDict[cpuDict['idnum']] = cpuDict
        CPUCount += 1
        cRow += 1

    outDict['admin'] = {}
    outDict['admin']['updated'] = (str)(datetime.now())[:-7]
    outDict['admin']['cpuCount'] = CPUCount
    outObj = json.dumps(outDict, indent = 4)
    outputFile = open((os.getcwd() + '/driverBin/data/CPUData.json'), "w")
    outputFile.write(outObj)
    outputFile.close()

# Read from internal previous human readable excel sheet holding data for each software
def readSoftwareList():
    SoftwareListFilepath = (os.getcwd() + '/../../../MASLD SOFTWARE LOGS.xlsx')
    softwareBook = load_workbook(SoftwareListFilepath, data_only = True)
    outDict = {}

    for sheetName in softwareBook.sheetnames:
        ws = softwareBook[sheetName]

        currSoftDict = {}
        currSoftDict['fields'] = [] # Keep track of the various fields for each software

        currCol = 66 # Starting the sheet at the ASCII value for a capitol B
        while(ws[f'{chr(currCol)}2'].value):
            currSoftDict['fields'].append(ws[f'{chr(currCol)}2'].value)
            currCol += 1

        row = 3
        while(ws[f'B{row}'].value):
            currCol = 66 # Starting the sheet at the ASCII value for a capitol B
            currSoftDict[ws[f'B{row}'].value] = {} # Declare a dictionary for each computer
            for field in currSoftDict['fields']:
                if(field == "PC NAME"):
                    currCol += 1
                    continue
                currSoftDict[ws[f'B{row}'].value][field] = ws[f'{chr(currCol)}{row}'].value
                currCol += 1
            row += 1

        currSoftDict['fields'].remove('PC NAME') # Remove the PC NAME field, it no longer is needed on the internal system
        outDict[sheetName] = currSoftDict
    
    outDict['admin'] = {}
    outDict['admin']['updated'] = (str)(datetime.now())[:-7]
    outObj = json.dumps(outDict, indent = 4)
    outputFile = open((os.getcwd() + '/driverBin/data/SoftwareData.json'), "w")
    outputFile.write(outObj)
    outputFile.close()