from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk

class EditCPUMenu(tk.Toplevel):

    def __init__(self, CPUMenu, currentCPU):
        super().__init__(CPUMenu)
        self.grab_set()
        self.focus()
        self.launchEditCPUDriver(currentCPU)

    # ==============================
    # INPUT CHECKING METHODS BELOW
    # ============================== 
    def checkInput(self, input):
        if((input == '') or (input == ' ')):
            return '--'
        else:
            return input

    # ==============================
    # SAVE METHODS BELOW
    # ============================== 
    def saveChanges(self, currentCPU, textList):
        # Save to excel sheet and currentCPU dictionary to see if it edits on back end.
        currentCPU["ipv4"] = self.checkInput(textList[0].get("1.0", 'end-1c'))
        currentCPU["location"] = self.checkInput(textList[1].get("1.0", 'end-1c'))
        currentCPU["userInit"] = self.checkInput(textList[2].get("1.0", 'end-1c'))
        currentCPU["user"] = self.checkInput(textList[3].get("1.0", 'end-1c'))
        currentCPU["build"] = self.checkInput(textList[4].get("1.0", 'end-1c'))
        currentCPU["type"] = self.checkInput(textList[5].get("1.0", 'end-1c'))
        currentCPU["cost"] = self.checkInput(textList[6].get("1.0", 'end-1c'))
        currentCPU["year"] = self.checkInput(textList[7].get("1.0", 'end-1c'))
        currentCPU["manufacturer"] = self.checkInput(textList[8].get("1.0", 'end-1c'))
        currentCPU["softwares"] = self.checkInput(textList[9].get("1.0", 'end-1c'))
        currentCPU["wifi"] = self.checkInput(textList[10].get("1.0", 'end-1c'))
        currentCPU["gpu"] = self.checkInput(textList[11].get("1.0", 'end-1c'))
        currentCPU["cpu"] = self.checkInput(textList[12].get("1.0", 'end-1c'))
        currentCPU["mobo"] = self.checkInput(textList[13].get("1.0", 'end-1c'))
        currentCPU["ram"] = self.checkInput(textList[14].get("1.0", 'end-1c'))
        currentCPU["storage"] = self.checkInput(textList[15].get("1.0", 'end-1c'))
        currentCPU["os"] = self.checkInput(textList[16].get("1.0", 'end-1c'))
        currentCPU["lastRowUpdate"] = (str)(datetime.now())[:-7]

        # Save changes to excel sheet.
        # TODO: Possibly shift off of this function and migrate to another location/file for optimization.
        try:
            CPUlistFile = load_workbook(os.getcwd() + '/../../../MASLD COMPUTER LOGS.xlsx', data_only = True)
            ws = CPUlistFile.active

            cRow = 3
            while (ws[f'B{cRow}'].value != currentCPU['idnum']):
                cRow += 1
            
            ws[f'C{cRow}'] = currentCPU["ipv4"]
            ws[f'D{cRow}'] = currentCPU["location"]
            ws[f'E{cRow}'] = currentCPU["userInit"]
            ws[f'F{cRow}'] = currentCPU["user"]
            ws[f'G{cRow}'] = currentCPU["build"]
            ws[f'H{cRow}'] = currentCPU["type"]
            ws[f'I{cRow}'] = currentCPU["cost"]
            ws[f'J{cRow}'] = currentCPU["year"]
            ws[f'K{cRow}'] = currentCPU["manufacturer"]
            ws[f'L{cRow}'] = currentCPU["softwares"]
            ws[f'M{cRow}'] = currentCPU["wifi"]
            ws[f'N{cRow}'] = currentCPU["gpu"]
            ws[f'O{cRow}'] = currentCPU["cpu"]
            ws[f'P{cRow}'] = currentCPU["mobo"]
            ws[f'Q{cRow}'] = currentCPU["ram"]
            ws[f'R{cRow}'] = currentCPU["storage"]
            ws[f'S{cRow}'] = currentCPU["os"]
            ws[f'T{cRow}'] = currentCPU["lastRowUpdate"]

            CPUlistFile.save(filename = (os.getcwd() + '/../../../MASLD COMPUTER LOGS.xlsx'))
            CPUlistFile.close()
        except Exception as E:
            tk.messagebox.showerror(title = " FILE OPEN ERROR", message = "Failiure in opening/saving Excel workbook containing CPU list. Changes have not been saved.")
            print(E)

        self.grab_release()
        self.destroy()

    # ==============================
    # RESET TEXT METHODS BELOW
    # ============================== 
    def resetText(self, textList: list[tk.Text]):
        # Ask user if they are sure they want to reset changes
        if(not(tk.messagebox.askyesno(" Warning", "This action will reset any unsaved changes you have made to this worksheet. Are you sure?"))):
            return

        # Remove text currently in the text element
        textList[0].delete("1.0", tk.END)
        # Replace with original text from inital load of window
        textList[0].insert(tk.END, self.ip_var.get())

        # Above comments are repeated throughout the end of this function
        textList[1].delete("1.0", tk.END)
        textList[1].insert(tk.END, self.loc_var.get())

        textList[2].delete("1.0", tk.END)
        textList[2].insert(tk.END, self.mui_var.get())

        textList[3].delete("1.0", tk.END)
        textList[3].insert(tk.END, self.mu_var.get())

        textList[4].delete("1.0", tk.END)
        textList[4].insert(tk.END, self.build_var.get())

        textList[5].delete("1.0", tk.END)
        textList[5].insert(tk.END, self.type_var.get())

        textList[6].delete("1.0", tk.END)
        textList[6].insert(tk.END, self.cost_var.get())

        textList[7].delete("1.0", tk.END)
        textList[7].insert(tk.END, self.year_var.get())

        textList[8].delete("1.0", tk.END)
        textList[8].insert(tk.END, self.manu_var.get())

        textList[9].delete("1.0", tk.END)
        textList[9].insert(tk.END, self.soft_var.get())

        textList[10].delete("1.0", tk.END)
        textList[10].insert(tk.END, self.internet_var.get())

        textList[11].delete("1.0", tk.END)
        textList[11].insert(tk.END, self.gpu_var.get())

        textList[12].delete("1.0", tk.END)
        textList[12].insert(tk.END, self.cpu_var.get())

        textList[13].delete("1.0", tk.END)
        textList[13].insert(tk.END, self.mobo_var.get())

        textList[14].delete("1.0", tk.END)
        textList[14].insert(tk.END, self.ram_var.get())

        textList[15].delete("1.0", tk.END)
        textList[15].insert(tk.END, self.storage_var.get())

        textList[16].delete("1.0", tk.END)
        textList[16].insert(tk.END, self.os_var.get())


    # ==============================
    # TERMINATE METHODS BELOW
    # ============================== 
    def terminate(self):
        response = tk.messagebox.askokcancel(parent = self, title = " Are you sure?", message = "Any unsaved changes will be lost.\nAre you sure you want to cancel?")
        if(response):
            self.grab_release()
            self.destroy()
        else:
            return

    # ==============================
    # DRIVER
    # ==============================
    def launchEditCPUDriver(self, currentCPU):
        self.title(f' Edit Data for {currentCPU["idnum"]} - InvenPro')

        # Create label on top of window, serve as direction window.
        self.eDirectionLabel = tk.Label(master = self, bg = 'grey', text = f'Editing data for {currentCPU["idnum"]}, click \"Apply\" to save them.\nChanges are FINAL.', font = 'bold', borderwidth = 2)
        self.eDirectionLabel.grid(row = 0, column = 0, columnspan = 3, sticky = "EW")

        # Create center canvas, required for scrollbar to function as needed.
        eCenterCanvas = tk.Canvas(master = self, height = 400, width = 500, highlightthickness = 0)
        eCenterCanvas.grid(row = 1, column = 0, columnspan = 3, sticky = "NSEW")

        # Create frame to shove into the canvas, frame is not placed in canvas yet for population purposes
        eDataConFrame = tk.Frame(master = eCenterCanvas)

        eCenterCanvas.create_window((0,0), window = eDataConFrame, anchor = 'nw')

        # Declare scrollbar for canvas
        vertScrollBar = tk.Scrollbar(master = self, orient = tk.VERTICAL, command = eCenterCanvas.yview)
        vertScrollBar.grid(row = 1, column = 2, sticky = "NSE")
        
        eCenterCanvas.config(yscrollcommand = vertScrollBar.set)

        # Strange line of code, required for scrollbar to work, defines scrollable area
        eDataConFrame.bind('<Configure>', lambda event: eCenterCanvas.config(scrollregion = eCenterCanvas.bbox('all')))

        textList = [] # List containing all of the text elements from the screen, for later parsing.

        self.ip_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["ipv4"]}')
        self.loc_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["location"]}')
        self.mui_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["userInit"]}')
        self.mu_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["user"]}')
        self.build_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["build"]}')
        self.type_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["type"]}')
        self.cost_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["cost"]}')
        self.year_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["year"]}')
        self.manu_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["manufacturer"]}')
        self.soft_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["softwares"]}')
        self.internet_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["wifi"]}')
        self.gpu_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["gpu"]}')
        self.cpu_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["cpu"]}')
        self.mobo_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["mobo"]}')
        self.ram_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["ram"]}')
        self.storage_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["storage"]}')
        self.os_var = tk.StringVar(master = eDataConFrame, value = f'{currentCPU["os"]}')

        cIPLabel = tk.Label(master = eDataConFrame, text = f'IPV4 ADDRESS:', font = 'Times 11')
        cIPDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cIPDataEntry.insert(tk.END, self.ip_var.get())
        textList.append(cIPDataEntry) # textList Index 0
        cIPLabel.grid(row = 1, column = 0)
        cIPDataEntry.grid(row = 1, column = 1, sticky = "E")

        cLocLabel = tk.Label(master = eDataConFrame, text = f'LOCATION:', font = 'Times 11')
        cLocDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cLocDataEntry.insert(tk.END, self.loc_var.get())
        textList.append(cLocDataEntry) # textList Index 1
        cLocLabel.grid(row = 2, column = 0)
        cLocDataEntry.grid(row = 2, column = 1, sticky = "E")
        
        cMUILabel = tk.Label(master = eDataConFrame, text = f'MAIN USER INITALS:', font = 'Times 11')
        cMUIDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cMUIDataEntry.insert(tk.END, self.mui_var.get())
        textList.append(cMUIDataEntry) # textList Index 2
        cMUILabel.grid(row = 3, column = 0)
        cMUIDataEntry.grid(row = 3, column = 1, sticky = "E")

        cMULabel = tk.Label(master = eDataConFrame, text = f'MAIN USER:', font = 'Times 11')
        cMUDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cMUDataEntry.insert(tk.END, self.mu_var.get())
        textList.append(cMUDataEntry) # textList Index 3
        cMULabel.grid(row = 4, column = 0)
        cMUDataEntry.grid(row = 4, column = 1, sticky = "E")

        cBuildLabel = tk.Label(master = eDataConFrame, text = f'BUILD:', font = 'Times 11')
        cBuildDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cBuildDataEntry.insert(tk.END, self.build_var.get())
        textList.append(cBuildDataEntry) # textList Index 4
        cBuildLabel.grid(row = 5, column = 0)
        cBuildDataEntry.grid(row = 5, column = 1, sticky = "E")

        cTypeLabel = tk.Label(master = eDataConFrame, text = f'TYPE:', font = 'Times 11')
        cTypeDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cTypeDataEntry.insert(tk.END, self.type_var.get())
        textList.append(cTypeDataEntry) # textList Index 5
        cTypeLabel.grid(row = 6, column = 0)
        cTypeDataEntry.grid(row = 6, column = 1, sticky = "E")

        cCostLabel = tk.Label(master = eDataConFrame, text = f'COST:', font = 'Times 11')
        cCostDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cCostDataEntry.insert(tk.END, self.cost_var.get())
        textList.append(cCostDataEntry) # textList Index 6
        cCostLabel.grid(row = 7, column = 0)
        cCostDataEntry.grid(row = 7, column = 1, sticky = "E")

        cYearLabel = tk.Label(master = eDataConFrame, text = f'YEAR:', font = 'Times 11')
        cYearDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cYearDataEntry.insert(tk.END, self.year_var.get())
        textList.append(cYearDataEntry) # textList Index 7
        cYearLabel.grid(row = 8, column = 0)
        cYearDataEntry.grid(row = 8, column = 1, sticky = "E")

        cManuLabel = tk.Label(master = eDataConFrame, text = f'MANUFACTURER:', font = 'Times 11')
        cManuDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cManuDataEntry.insert(tk.END, self.manu_var.get())
        textList.append(cManuDataEntry) # textList Index 8
        cManuLabel.grid(row = 9, column = 0)
        cManuDataEntry.grid(row = 9, column = 1, sticky = "E")

        cSoftwaresLabel = tk.Label(master = eDataConFrame, text = f'SOFTWARES:', font = 'Times 11')
        cSoftwaresDataEntry = tk.Text(master = eDataConFrame, height = ((self.soft_var.get().count('\n')) + 1), width = 40)
        cSoftwaresDataEntry.insert(tk.END, self.soft_var.get())
        textList.append(cSoftwaresDataEntry) # textList Index 9
        cSoftwaresLabel.grid(row = 10, column = 0)
        cSoftwaresDataEntry.grid(row = 10, column = 1, sticky = "E")

        cInternetLabel = tk.Label(master = eDataConFrame, text = f'INTERNET:', font = 'Times 11')
        cInternetDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cInternetDataEntry.insert(tk.END, self.internet_var.get())
        textList.append(cInternetDataEntry) # textList Index 10
        cInternetLabel.grid(row = 11, column = 0)
        cInternetDataEntry.grid(row = 11, column = 1, sticky = "E")

        cGPULabel = tk.Label(master = eDataConFrame, text = f'GPU:', font = 'Times 11')
        cGPUDataEntry = tk.Text(master = eDataConFrame, height = ((self.gpu_var.get().count('\n')) + 1), width = 40)
        cGPUDataEntry.insert(tk.END, self.gpu_var.get())
        textList.append(cGPUDataEntry) # textList Index 11
        cGPULabel.grid(row = 12, column = 0)
        cGPUDataEntry.grid(row = 12, column = 1, sticky = "E")
        
        cCPULabel = tk.Label(master = eDataConFrame, text = f'CPU:', font = 'Times 11')
        cCPUDataEntry = tk.Text(master = eDataConFrame, height = ((self.cpu_var.get().count('\n')) + 1), width = 40)
        cCPUDataEntry.insert(tk.END, self.cpu_var.get())
        textList.append(cCPUDataEntry) # textList Index 12
        cCPULabel.grid(row = 13, column = 0)
        cCPUDataEntry.grid(row = 13, column = 1, sticky = "E")
        
        cMoboLabel = tk.Label(master = eDataConFrame, text = f'MOTHERBOARD:', font = 'Times 11')
        cMoboDataEntry = tk.Text(master = eDataConFrame, height = ((self.mobo_var.get().count('\n')) + 1), width = 40)
        cMoboDataEntry.insert(tk.END, self.mobo_var.get())
        textList.append(cMoboDataEntry) # textList Index 13
        cMoboLabel.grid(row = 14, column = 0)
        cMoboDataEntry.grid(row = 14, column = 1, sticky = "E")
        
        cRAMLabel = tk.Label(master = eDataConFrame, text = f'RAM:', font = 'Times 11')
        cRAMDataEntry = tk.Text(master = eDataConFrame, height = ((self.ram_var.get().count('\n')) + 1), width = 40)
        cRAMDataEntry.insert(tk.END, self.ram_var.get())
        textList.append(cRAMDataEntry) # textList Index 14
        cRAMLabel.grid(row = 15, column = 0)
        cRAMDataEntry.grid(row = 15, column = 1, sticky = "E")
        
        cStorageLabel = tk.Label(master = eDataConFrame, text = f'STORAGE:', font = 'Times 11')
        cStorageDataEntry = tk.Text(master = eDataConFrame, height = ((self.storage_var.get().count('\n')) + 1), width = 40)
        cStorageDataEntry.insert(tk.END, self.storage_var.get())
        textList.append(cStorageDataEntry) # textList Index 15
        cStorageLabel.grid(row = 16, column = 0)
        cStorageDataEntry.grid(row = 16, column = 1, sticky = "E")
        
        cOSLabel = tk.Label(master = eDataConFrame, text = f'OPERATING SYSTEM:', font = 'Times 11')
        cOSDataEntry = tk.Text(master = eDataConFrame, height = 1, width = 40)
        cOSDataEntry.insert(tk.END, self.os_var.get())
        textList.append(cOSDataEntry) # textList Index 16
        cOSLabel.grid(row = 17, column = 0)
        cOSDataEntry.grid(row = 17, column = 1, sticky = "E")

        # Create master frame for button organization
        cBttnFrame = tk.Frame(master = self)
        # Format columns of master button frame
        cBttnFrame.grid_columnconfigure(0, weight = 1)
        cBttnFrame.grid_columnconfigure(1, minsize = 260)
        cBttnFrame.grid_columnconfigure(2, weight = 1)
        cBttnFrame.grid_columnconfigure(3, weight = 1)
        # Create Save, Cancel, and Reset Button
        cResetButtn = ttk.Button(master = cBttnFrame, text = "Reset", style = "M.TButton", command = lambda: self.resetText(textList))
        cSaveButtn = ttk.Button(master = cBttnFrame, text = "Apply", style = "M.TButton", command = lambda: self.saveChanges(currentCPU, textList))
        cCancelButtn = ttk.Button(master = cBttnFrame, text = "Cancel", style = "M.TButton", command = lambda: self.terminate())
        # Place Save, Cancel, and Reset Button
        cSaveButtn.grid(row = 0, column = 0, sticky = "E", padx = .5, pady = .5)
        cResetButtn.grid(row = 0, column = 2, sticky = "E", padx = .5, pady = .5)
        cCancelButtn.grid(row = 0, column = 3, sticky = "E", padx = .5, pady = .5)

        # Place button frame
        cBttnFrame.grid(row = 2, column = 2, sticky = "EW")

        # Intercept close button
        self.protocol("WM_DELETE_WINDOW", lambda: self.terminate())

        self.resizable(0,0)
        # self.mainloop() -- Commented out to see if this would fix wait_window() calls by parent process, UDPATE: It worked.
        return