from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk
from .InputDialogue import InputDioBox
from ..utils import settings as GLOBAL

class AddCPUMenu(tk.Toplevel):
    TEXTWIDTH = 40

    def __init__(self, CPUMenu):
        super().__init__(CPUMenu) # Call super constructor
        self.grab_set()
        self.focus_set()
        self.buildAddWindow()
    
    # ==============================
    # HELPER METHODS BELOW
    # ==============================

    # Provides the next open ID using the Json and adding one to the CPU count
    def getNextOpenID(self):
        self.aIDButton.config(state = 'disabled') #Disable button while function runs

        # Determine next open number for PC
        nextNum = (GLOBAL.pcDict['admin']['cpuCount'] + 1)
        if(nextNum < 10):
            out = f'MASLD-PC-00{nextNum}'
        elif(nextNum < 100):
            out = f'MASLD-PC-0{nextNum}'
        else:
            out = f'MASLD-PC-{nextNum}'
        
        self.aIDText.delete("1.0", tk.END) # Clear textbox incase it has already been filled out
        self.aIDText.insert(tk.END, out) # Insert new ID
        self.aIDButton.config(state = 'normal') # Reactivate button

    # Expands the passed in Text widget by one line, assuming the user has pressed enter
    def addTextLine(self, textBox: tk.Text):
        height = textBox.get("1.0", tk.END).count('\n')
        textBox.config(height = height + 1)

    # Runs checks to ensure the passed in ID number is acceptable and will work in the system. Extremely important. TODO: Optimize
    def checkIDNum(self, idnum):
        print("DEBUG: Checking IDNum in AddCPU.")
        out = tk.StringVar(master = self, value = idnum) # Use Stringvar to retain data after window close
        while(True):
            if(out.get() == "" or out.get() == None or out.get() == "\n"): # Check for empty IDs
                dialogWin = InputDioBox(self, " ERROR - IDNUM Invalid", "You MUST supply an ID number for the new PC.\nPlease supply a valid ID.", out) # Initalize and wait for popup.
                self.wait_window(dialogWin)
                if(out.get() == "INPUTDIOCANCEL.DEBUG.INVENPRO"): # Return none if user cancelled edit window
                    print("DEBUG: User Cancelled InputDialogue Popup in AddCPU.checkIDNum().")
                    return None
                else:
                    continue
            if(not("MASLD-PC-" in out.get())): # Check for incorrect ID format
                dialogWin = InputDioBox(self, " ERROR - IDNUM Invalid", "ID must be in the following format: \"MASLD-PC-XXX\".\nPlease enter valid ID", out) # Initalize and wait for popup.
                self.wait_window(dialogWin)
                if(out.get() == "INPUTDIOCANCEL.DEBUG.INVENPRO"): # Return none if user cancelled edit window
                    print("DEBUG: User Cancelled InputDialogue Popup in AddCPU.checkIDNum().") 
                    return None
                else:
                    continue
            if(out.get() in GLOBAL.pcDict): # Check for duplicate IDs
                dialogWin = InputDioBox(self, " ERROR - IDNUM Invalid", "The ID you supplied has already been taken.\nPlease input an ID that is not taken.", out) # Initalize and wait for popup.
                self.wait_window(dialogWin)
                if(out.get() == "INPUTDIOCANCEL.DEBUG.INVENPRO"): # Return none if user cancelled edit window
                    print("DEBUG: User Cancelled InputDialogue Popup in AddCPU.checkIDNum().")
                    return None
                else:
                    continue
            break
        return out.get()

    # Replaces all empty inputs with "--"
    def checkEmptyInput(self, text, field):
        out = tk.StringVar(master = self, value = text) # Use Stringvar to retain data after window close
        if((text == "") or (text == None) or (text == "\n") or (text == " ")): # Check text for empty value
            out.set(value = "--") # Replace empty value with "--"
        return out.get()

    # Check all input fields and format data pulled from them.
    def parseAndCheckTexts(self):
        # Parse and save all inputs into a dict
        outDict = {}
        # Parse and Check ID Number
        outID = self.checkIDNum(self.aIDText.get("1.0", 'end-1c'))
        if(outID == None): # Return none if user cancelled during edit phase
            return None
        outDict["idnum"] = outID
        # Parse and Check Remainder of fields
        outDict["ipv4"] = self.checkEmptyInput(self.aIPText.get("1.0", 'end-1c'), "IPV4")
        outDict["location"] = self.checkEmptyInput(self.aLocationText.get("1.0", 'end-1c'), "LOCATION")
        outDict["userInit"] = self.checkEmptyInput(self.aUserInitText.get("1.0", 'end-1c'), "MAIN USER INITALS")
        outDict["user"] = self.checkEmptyInput(self.aUserText.get("1.0", 'end-1c'), "MAIN USER")
        outDict["build"] = self.checkEmptyInput(self.aBuildText.get("1.0", 'end-1c'), "BUILD")
        outDict["type"] = self.checkEmptyInput(self.aTypeText.get("1.0", 'end-1c'), "TYPE")
        outDict["cost"] = self.checkEmptyInput(self.aCostText.get("1.0", 'end-1c'), "COST")
        outDict["year"] = self.checkEmptyInput(self.aYearText.get("1.0", 'end-1c'), "YEAR")
        outDict["manufacturer"] = self.checkEmptyInput(self.aManufacturerText.get("1.0", 'end-1c'), "MANUFACTURER")
        outDict["softwares"] = self.checkEmptyInput(self.aSoftwaresText.get("1.0", 'end-1c'), "SOFTWARES")
        outDict["wifi"] = self.checkEmptyInput(self.aInternetText.get("1.0", 'end-1c'), "INTERNET")
        outDict["gpu"] = self.checkEmptyInput(self.aGPUText.get("1.0", 'end-1c'), "GPU")
        outDict["cpu"] = self.checkEmptyInput(self.aCPUText.get("1.0", 'end-1c'), "CPU")
        outDict["mobo"] = self.checkEmptyInput(self.aMotherboardText.get("1.0", 'end-1c'), "MOTHERBOARD")
        outDict["ram"] = self.checkEmptyInput(self.aRAMText.get("1.0", 'end-1c'), "RAM")
        outDict["storage"] = self.checkEmptyInput(self.aStorageText.get("1.0", 'end-1c'), "STORAGE")
        outDict["os"] = self.checkEmptyInput(self.aOSText.get("1.0", 'end-1c'), "OS")
        # Update time also serves as time of addition to system
        outDict["lastRowUpdate"] = (str)(datetime.now())[:-7]

        return outDict

    # ==============================
    # SAVE METHOD BELOW
    # ==============================
    def savePC(self):
        newPC = self.parseAndCheckTexts()
        if(newPC == None):
            tk.messagebox.showwarning(parent = self, title = " Attention", message = "Addition of new PC cancelled. PC has not been added to system. You may retry or cancel addition of new PC.")
            return
        GLOBAL.pcDict[newPC['idnum']] = newPC
        GLOBAL.pcDict['admin']['cpuCount'] += 1

        # Save newPC to excel sheet.
        # TODO: Possibly shift off of this function and migrate to another location/file for optimization.
        try:
            CPUlistFile = load_workbook(os.getcwd() + '/../../../MASLD COMPUTER LOGS.xlsx', data_only = True)
            ws = CPUlistFile.active

            cRow = 3
            while (ws[f'B{cRow}'].value):
                cRow += 1
            
            ws[f'B{cRow}'] = newPC["idnum"]
            ws[f'C{cRow}'] = newPC["ipv4"]
            ws[f'D{cRow}'] = newPC["location"]
            ws[f'E{cRow}'] = newPC["userInit"]
            ws[f'F{cRow}'] = newPC["user"]
            ws[f'G{cRow}'] = newPC["build"]
            ws[f'H{cRow}'] = newPC["type"]
            ws[f'I{cRow}'] = newPC["cost"]
            ws[f'J{cRow}'] = newPC["year"]
            ws[f'K{cRow}'] = newPC["manufacturer"]
            ws[f'L{cRow}'] = newPC["softwares"]
            ws[f'M{cRow}'] = newPC["wifi"]
            ws[f'N{cRow}'] = newPC["gpu"]
            ws[f'O{cRow}'] = newPC["cpu"]
            ws[f'P{cRow}'] = newPC["mobo"]
            ws[f'Q{cRow}'] = newPC["ram"]
            ws[f'R{cRow}'] = newPC["storage"]
            ws[f'S{cRow}'] = newPC["os"]
            ws[f'T{cRow}'] = newPC["lastRowUpdate"]

            CPUlistFile.save(filename = (os.getcwd() + '/../../../MASLD COMPUTER LOGS.xlsx'))
            CPUlistFile.close()
        except Exception as E:
            tk.messagebox.showerror(title = " FILE OPEN ERROR", message = "Failiure in opening/saving Excel workbook containing CPU list. New PC has not been saved.")
            GLOBAL.pcDict.pop(newPC['idnum'], None)
            GLOBAL.pcDict['admin']['cpuCount'] -= 1
            print(E)


        self.grab_release()
        self.destroy()

    # ==============================
    # TERMINATE METHOD BELOW
    # ==============================
    def terminate(self):
        response = tk.messagebox.askokcancel(parent = self, title = " Are you sure?", message = "Current PC addition will be lost.\nAre you sure you want to cancel?")
        if(response):
            self.grab_release()
            self.destroy()
        else:
            return

    # ==============================
    # DRIVER
    # ==============================
    def buildAddWindow(self):
        self.title(f' Add PC - InvenPro')

        # Create and place label for directions
        aDirectionLabel = tk.Label(master = self, text = "Fill Out Information to Add new PC", font = 'Times 20 bold', bg = 'grey')
        aDirectionLabel.grid(row = 0, column = 0, columnspan = 3, sticky = "ESW")

        # Create center canvas, required for scrollbar to function as needed.
        aCenterCanvas = tk.Canvas(master = self, height = 400, width = 500, highlightthickness = 0)
        aCenterCanvas.grid(row = 1, column = 0, columnspan = 3, sticky = "NSEW")

        # Create frame for labels and text boxes
        aDataFrame = tk.Frame(master = aCenterCanvas)

        # Attach frame to internals of center canvas
        aCenterCanvas.create_window((0,0), window = aDataFrame, anchor = 'nw')

        # Create and place Scrollbar
        aVertScrollBar = tk.Scrollbar(master = self, orient = tk.VERTICAL, relief = 'sunken', bd = 2, command = aCenterCanvas.yview)
        aVertScrollBar.grid(row = 1, column = 2, sticky = "NSE")

        # Config scrollbar to have better relation to canvas in size
        aCenterCanvas.config(yscrollcommand = aVertScrollBar.set)

        # Strange line of code, required for scrollbar to work, defines scrollable area
        aDataFrame.bind('<Configure>', lambda event: aCenterCanvas.config(scrollregion = aCenterCanvas.bbox('all')))

        # Declare ID Label, Text box, and 'Next ID' Button
        aIDLabel = tk.Label(master = aDataFrame, text = "MASLD ID:", font = 'Times 11')
        self.aIDText = tk.Text(master = aDataFrame, height = 1, width = 30)
        self.aIDButton = ttk.Button(master = aDataFrame, text = "Next Open ID", style = "M.TButton", command = lambda: self.getNextOpenID())
        # Place ID Label, Text box, and 'Next ID' Button
        aIDLabel.grid(row = 0, column = 0, padx = 1)
        self.aIDText.grid(row = 0, column = 1, sticky = "E")
        self.aIDButton.grid(row = 0, column = 2, sticky = "W")

        # Declare IP Label and Text box
        aIPLabel = tk.Label(master = aDataFrame, text = "IPV4:", font = 'Times 11')
        self.aIPText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place IP Label and Text box
        aIPLabel.grid(row = 1, column = 0, padx = 1)
        self.aIPText.grid(row = 1, column = 1, columnspan = 2, sticky = "W")

        # Declare Location Label and Text box
        aLocationLabel = tk.Label(master = aDataFrame, text = "LOCATION:", font = 'Times 11')
        self.aLocationText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place Location Label and Text box
        aLocationLabel.grid(row = 2, column = 0, padx = 1)
        self.aLocationText.grid(row = 2, column = 1, columnspan = 2, sticky = "W")

        # Declare User Initals Label and Text box
        aUserInitLabel = tk.Label(master = aDataFrame, text = "MAIN USER INITALS:", font = 'Times 11')
        self.aUserInitText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place User Initals Label and Text box
        aUserInitLabel.grid(row = 3, column = 0, padx = 1)
        self.aUserInitText.grid(row = 3, column = 1, columnspan = 2, sticky = "W")

        # Declare User Name Label and Text box
        aUserLabel = tk.Label(master = aDataFrame, text = "MAIN USER:", font = 'Times 11')
        self.aUserText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place User Name Label and Text box
        aUserLabel.grid(row = 4, column = 0, padx = 1)
        self.aUserText.grid(row = 4, column = 1, columnspan = 2, sticky = "W")

        # Declare Build Name Label and Text box
        aBuildLabel = tk.Label(master = aDataFrame, text = "BUILD:", font = 'Times 11')
        self.aBuildText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place Build Name Label and Text box
        aBuildLabel.grid(row = 5, column = 0, padx = 1)
        self.aBuildText.grid(row = 5, column = 1, columnspan = 2, sticky = "W")

        # Declare Type Name Label and Text box
        aTypeLabel = tk.Label(master = aDataFrame, text = "TYPE:", font = 'Times 11')
        self.aTypeText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place Type Name Label and Text box
        aTypeLabel.grid(row = 6, column = 0, padx = 1)
        self.aTypeText.grid(row = 6, column = 1, columnspan = 2, sticky = "W")

        # Declare Cost Name Label and Text box
        aCostLabel = tk.Label(master = aDataFrame, text = "COST:", font = 'Times 11')
        self.aCostText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place Cost Name Label and Text box
        aCostLabel.grid(row = 7, column = 0, padx = 1)
        self.aCostText.grid(row = 7, column = 1, columnspan = 2, sticky = "W")

        # Declare Year Name Label and Text box
        aYearLabel = tk.Label(master = aDataFrame, text = "YEAR:", font = 'Times 11')
        self.aYearText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place Year Name Label and Text box
        aYearLabel.grid(row = 8, column = 0, padx = 1)
        self.aYearText.grid(row = 8, column = 1, columnspan = 2, sticky = "W")

        # Declare Manufacturer Name Label and Text box
        aManufacturerLabel = tk.Label(master = aDataFrame, text = "MANUFACTURER:", font = 'Times 11')
        self.aManufacturerText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place Manufacturer Name Label and Text box
        aManufacturerLabel.grid(row = 9, column = 0, padx = 1)
        self.aManufacturerText.grid(row = 9, column = 1, columnspan = 2, sticky = "W")

        # Declare Softwares Name Label, Text box, and ADDLINE Button
        aSoftwaresLabel = tk.Label(master = aDataFrame, text = "SOFTWARES:", font = 'Times 11')
        self.aSoftwaresText = tk.Text(master = aDataFrame, height = 1, width = 30)
        aSoftAddLineButtn = ttk.Button(master = aDataFrame, text = "Add Line", command = lambda: self.addTextLine(self.aSoftwaresText))
        # Place Softwares Name Label, Text box, and ADDLINE Button
        aSoftwaresLabel.grid(row = 10, column = 0, padx = 1)
        self.aSoftwaresText.grid(row = 10, column = 1, sticky = "W")
        aSoftAddLineButtn.grid(row = 10, column = 2, padx = 0.5, sticky = "E")

        # Declare Internet Name Label and Text box
        aInternetLabel = tk.Label(master = aDataFrame, text = "INTERNET:", font = 'Times 11')
        self.aInternetText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place Internet Name Label and Text box
        aInternetLabel.grid(row = 11, column = 0, padx = 1)
        self.aInternetText.grid(row = 11, column = 1, columnspan = 2, sticky = "W")

        # Declare GPU Name Label, Text box, and ADDLINE Button
        aGPULabel = tk.Label(master = aDataFrame, text = "GPU:", font = 'Times 11')
        self.aGPUText = tk.Text(master = aDataFrame, height = 1, width = 30)
        aGPUAddLineButtn = ttk.Button(master = aDataFrame, text = "Add Line", command = lambda: self.addTextLine(self.aGPUText))
        # Place GPU Name Label, Text box, and ADDLINE Button
        aGPULabel.grid(row = 12, column = 0, padx = 1)
        self.aGPUText.grid(row = 12, column = 1, sticky = "W")
        aGPUAddLineButtn.grid(row = 12, column = 2, padx = 0.5, sticky = "E")

        # Declare CPU Name Label, Text box, and ADDLINE Button
        aCPULabel = tk.Label(master = aDataFrame, text = "CPU:", font = 'Times 11')
        self.aCPUText = tk.Text(master = aDataFrame, height = 1, width = 30)
        aCPUAddLineButtn = ttk.Button(master = aDataFrame, text = "Add Line", command = lambda: self.addTextLine(self.aCPUText))
        # Place CPU Name Label, Text box, and ADDLINE Button
        aCPULabel.grid(row = 13, column = 0, padx = 1)
        self.aCPUText.grid(row = 13, column = 1, sticky = "W")
        aCPUAddLineButtn.grid(row = 13, column = 2, padx = 0.5, sticky = "E")

        # Declare Motherboard Name Label and Text box
        aMotherboardLabel = tk.Label(master = aDataFrame, text = "MOTHERBOARD:", font = 'Times 11')
        self.aMotherboardText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place Motherboard Name Label and Text box
        aMotherboardLabel.grid(row = 14, column = 0, padx = 1)
        self.aMotherboardText.grid(row = 14, column = 1, columnspan = 2, sticky = "W")

        # Declare RAM Name Label, Text box, and ADDLINE Button
        aRAMLabel = tk.Label(master = aDataFrame, text = "RAM:", font = 'Times 11')
        self.aRAMText = tk.Text(master = aDataFrame, height = 1, width = 30)
        aRAMAddLineButtn = ttk.Button(master = aDataFrame, text = "Add Line", command = lambda: self.addTextLine(self.aRAMText))
        # Place RAM Name Label, Text box, and ADDLINE Button
        aRAMLabel.grid(row = 15, column = 0, padx = 1)
        self.aRAMText.grid(row = 15, column = 1, sticky = "W")
        aRAMAddLineButtn.grid(row = 15, column = 2, padx = 0.5, sticky = "E")

        # Declare Storage Name Label, Text box, and ADDLINE Button
        aStorageLabel = tk.Label(master = aDataFrame, text = "STORAGE:", font = 'Times 11')
        self.aStorageText = tk.Text(master = aDataFrame, height = 1, width = 30)
        aStorageAddLineButtn = ttk.Button(master = aDataFrame, text = "Add Line", command = lambda: self.addTextLine(self.aStorageText))
        # Place Storage Name Label, Text box, and ADDLINE Button
        aStorageLabel.grid(row = 16, column = 0, padx = 1)
        self.aStorageText.grid(row = 16, column = 1, sticky = "W")
        aStorageAddLineButtn.grid(row = 16, column = 2, padx = 0.5, sticky = "E")

        # Declare OS Name Label and Text box
        aOSLabel = tk.Label(master = aDataFrame, text = "OS:", font = 'Times 11')
        self.aOSText = tk.Text(master = aDataFrame, height = 1, width = self.TEXTWIDTH)
        # Place OS Name Label and Text box
        aOSLabel.grid(row = 17, column = 0, padx = 1)
        self.aOSText.grid(row = 17, column = 1, columnspan = 2, sticky = "W")

        # Create subFrame for Button placements
        aButtnFrame = tk.Frame(master = self)

        # Create Save and Cancel Button
        aSaveButtn = ttk.Button(master = aButtnFrame, text = "Save New PC", style = "M.TButton", command = lambda: self.savePC())
        aCancelButtn = ttk.Button(master = aButtnFrame, text = "Cancel", style = "M.TButton", command = lambda: self.terminate())
        # Place Save and Cancel Button
        aSaveButtn.grid(row = 0, column = 0, padx = .5, pady = .5, sticky = "E")
        aCancelButtn.grid(row = 0, column = 1, padx = .5, pady = .5, sticky = "W")

        # Place button subFrame
        aButtnFrame.grid(row = 2, column = 2, sticky = "E")

        # Intercept close button
        self.protocol("WM_DELETE_WINDOW", lambda: self.terminate())

        self.resizable(0,0)
        # self.mainloop() -- Commented out to see if this would fix wait_window() calls by parent process, UDPATE: It worked.
        return